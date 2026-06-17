"""
Master Customer Data Merger
============================
Reads 5 CSV files, merges them by email address, adds analysis columns,
and writes a clean, colour-coded Excel workbook.
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule
import os

# ─────────────────────────────────────────────────────────────────
# STEP 1 – SET THE FOLDER PATH
# ─────────────────────────────────────────────────────────────────
# Think of this as telling Python where your files live, like giving
# someone a street address before asking them to go pick something up.

DATA_FOLDER = r"c:\Users\sdhakshinamoorthi\OneDrive - AllegroMicro\Allegro Summer Intern\Practice data\Data"
OUTPUT_FILE = r"c:\Users\sdhakshinamoorthi\OneDrive - AllegroMicro\Allegro Summer Intern\Practice data\master_customers.xlsx"

print("=" * 60)
print("MASTER CUSTOMER DATA MERGER")
print("=" * 60)

# ─────────────────────────────────────────────────────────────────
# STEP 2 – READ EACH CSV FILE
# ─────────────────────────────────────────────────────────────────
# A CSV file is just a plain text file where each row is a customer
# and each value is separated by a comma.  pandas reads it into a
# "DataFrame" – imagine a very smart spreadsheet inside Python.

print("\nSTEP 1: Reading all 5 CSV files...")

salesforce = pd.read_csv(os.path.join(DATA_FOLDER, "salesforce.csv"))
billing    = pd.read_csv(os.path.join(DATA_FOLDER, "billing.csv"))
support    = pd.read_csv(os.path.join(DATA_FOLDER, "support.csv"))
marketing  = pd.read_csv(os.path.join(DATA_FOLDER, "marketing.csv"))
forecast   = pd.read_csv(os.path.join(DATA_FOLDER, "forecast.csv"))

print(f"  salesforce.csv  -> {len(salesforce)} rows, columns: {list(salesforce.columns)}")
print(f"  billing.csv     -> {len(billing)} rows, columns: {list(billing.columns)}")
print(f"  support.csv     -> {len(support)} rows, columns: {list(support.columns)}")
print(f"  marketing.csv   -> {len(marketing)} rows, columns: {list(marketing.columns)}")
print(f"  forecast.csv    -> {len(forecast)} rows, columns: {list(forecast.columns)}")

# ─────────────────────────────────────────────────────────────────
# STEP 3 – STANDARDISE THE EMAIL COLUMN NAME
# ─────────────────────────────────────────────────────────────────
# Each file calls the email column something slightly different:
#   salesforce → "Contact_Email"
#   billing    → "Email"
#   support    → "Email_Address"
#   marketing  → "Email"
#   forecast   → "Customer_Email"
#
# We rename every one of them to a single name: "email_key".
# This is like giving everyone the same name tag at a party so
# Python knows they are all talking about the same thing.

print("\nSTEP 2: Standardising email column names across all files...")

salesforce = salesforce.rename(columns={"Contact_Email": "email_key"})
billing    = billing.rename(columns={"Email":           "email_key"})
support    = support.rename(columns={"Email_Address":   "email_key"})
marketing  = marketing.rename(columns={"Email":         "email_key"})
forecast   = forecast.rename(columns={"Customer_Email": "email_key"})

# Make every email lower-case so "Bob@ACME.com" and "bob@acme.com"
# are treated as the same person.
for df in [salesforce, billing, support, marketing, forecast]:
    df["email_key"] = df["email_key"].str.strip().str.lower()

print("  Done – all email columns renamed to 'email_key' and lowercased.")

# ─────────────────────────────────────────────────────────────────
# STEP 4 – PICK ONLY THE COLUMNS WE WANT TO KEEP FROM EACH FILE
# ─────────────────────────────────────────────────────────────────
# We do not need every column from every file – for example, we do
# not need the internal IDs like "SF-001" or "BILL-101".
# Think of this as highlighting only the useful cells in a
# spreadsheet and ignoring the rest.

print("\nSTEP 3: Selecting useful columns from each file...")

sf_clean   = salesforce[["email_key", "Company_Name", "Phone", "Industry", "Account_Owner"]]

bill_clean = billing[["email_key", "Annual_Revenue", "Payment_Status", "Contract_Start"]]

sup_clean  = support[["email_key", "Issue_Type", "Priority",
                       "Status", "Open_Date"]].rename(
                           columns={"Status": "Support_Status",
                                    "Open_Date": "Ticket_Open_Date"})

mktg_clean = marketing[["email_key", "Segment", "Last_Campaign_Sent",
                         "Email_Opens", "Subscribed"]]

fcast_clean = forecast[["email_key", "Q1_Forecast", "Q2_Forecast",
                         "Q3_Forecast", "Q4_Forecast",
                         "Total_Year_Forecast", "Confidence"]]

print("  Selected columns for each file.")

# ─────────────────────────────────────────────────────────────────
# STEP 5 – MERGE ALL FILES ON EMAIL
# ─────────────────────────────────────────────────────────────────
# A "merge" works like a VLOOKUP in Excel: it matches rows from
# two tables that share the same email address and glues them
# side-by-side into one wider row.
#
# We use how="left" which means:
#   "Start with salesforce as the master list.
#    For every customer there, pull in matching data from the other
#    files.  If a file has no matching row, leave those cells blank
#    rather than dropping the customer."

print("\nSTEP 4: Merging all 5 files on the email address...")

master = sf_clean.copy()
master = master.merge(bill_clean,  on="email_key", how="left")
master = master.merge(sup_clean,   on="email_key", how="left")
master = master.merge(mktg_clean,  on="email_key", how="left")
master = master.merge(fcast_clean, on="email_key", how="left")

print(f"  Merged successfully -> {len(master)} rows, {len(master.columns)} columns.")

# ─────────────────────────────────────────────────────────────────
# STEP 6 – RENAME EMAIL KEY TO A FRIENDLY NAME
# ─────────────────────────────────────────────────────────────────
master = master.rename(columns={"email_key": "Email"})

# ─────────────────────────────────────────────────────────────────
# STEP 7 – ADD FLAG: OVERDUE PAYMENT OR OPEN SUPPORT TICKET
# ─────────────────────────────────────────────────────────────────
# We create a new TRUE/FALSE column called "Needs_Attention".
# It is TRUE when EITHER of the following is true:
#   • Payment_Status is "Overdue"  (the customer owes us money)
#   • Support_Status is "Open" or "In Progress"  (unresolved issue)
#
# Think of it like a red flag on a file folder – easy to spot at a
# glance which customers need follow-up.

print("\nSTEP 5: Adding 'Needs_Attention' flag column...")

overdue_payment   = master["Payment_Status"]  == "Overdue"
open_ticket       = master["Support_Status"].isin(["Open", "In Progress"])

master["Needs_Attention"] = overdue_payment | open_ticket   # "|" means OR

# Human-readable version (Yes / No instead of True / False)
master["Needs_Attention"] = master["Needs_Attention"].map({True: "Yes", False: "No"})

flagged = (master["Needs_Attention"] == "Yes").sum()
print(f"  {flagged} customer(s) flagged as needing attention.")

# ─────────────────────────────────────────────────────────────────
# STEP 8 – CONFIRM TOTAL YEAR FORECAST COLUMN
# ─────────────────────────────────────────────────────────────────
# The forecast file already has a "Total_Year_Forecast" column
# (Q1 + Q2 + Q3 + Q4 pre-summed).  We just rename it to something
# friendlier for the Excel output.

print("\nSTEP 6: Confirming Total Year Forecast Revenue column...")
master = master.rename(columns={"Total_Year_Forecast": "Total_Forecast_Revenue"})
print("  Column confirmed.")

# ─────────────────────────────────────────────────────────────────
# STEP 9 – SORT BY ANNUAL REVENUE (HIGHEST FIRST)
# ─────────────────────────────────────────────────────────────────
# Sorting puts the most valuable customers at the top of the list,
# like sorting a leaderboard from #1 downward.

print("\nSTEP 7: Sorting customers by Annual Revenue (highest first)...")
master = master.sort_values("Annual_Revenue", ascending=False).reset_index(drop=True)
print("  Sorted.")

# ─────────────────────────────────────────────────────────────────
# STEP 10 – REORDER COLUMNS SO THE MOST IMPORTANT ARE FIRST
# ─────────────────────────────────────────────────────────────────
column_order = [
    "Company_Name", "Email", "Industry", "Account_Owner",
    "Needs_Attention",
    "Payment_Status", "Annual_Revenue", "Contract_Start",
    "Total_Forecast_Revenue", "Q1_Forecast", "Q2_Forecast",
    "Q3_Forecast", "Q4_Forecast", "Confidence",
    "Support_Status", "Issue_Type", "Priority", "Ticket_Open_Date",
    "Segment", "Last_Campaign_Sent", "Email_Opens", "Subscribed",
    "Phone",
]
master = master[column_order]

# ─────────────────────────────────────────────────────────────────
# STEP 11 – WRITE TO EXCEL (raw data first, then format it)
# ─────────────────────────────────────────────────────────────────
print(f"\nSTEP 8: Writing data to Excel file...")

with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
    master.to_excel(writer, sheet_name="Master Customers", index=False)

print(f"  Raw data written.")

# ─────────────────────────────────────────────────────────────────
# STEP 12 – STYLE THE EXCEL FILE
# ─────────────────────────────────────────────────────────────────
# Now we open the file again and apply formatting – colours, bold
# headers, borders, number formatting – to make it easy to read.

print("\nSTEP 9: Applying Excel formatting (colours, borders, fonts)...")

wb = load_workbook(OUTPUT_FILE)
ws = wb["Master Customers"]

# ── Colour palette ──────────────────────────────────────────────
HEADER_BG   = PatternFill("solid", fgColor="1F3864")   # dark navy
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)

FLAG_YES_BG = PatternFill("solid", fgColor="FFD7D7")   # soft red
FLAG_NO_BG  = PatternFill("solid", fgColor="D7F0D7")   # soft green

ALT_ROW_BG  = PatternFill("solid", fgColor="EEF2F7")   # light blue-grey
WHITE_BG    = PatternFill("solid", fgColor="FFFFFF")

BORDER_SIDE = Side(style="thin", color="CCCCCC")
THIN_BORDER = Border(left=BORDER_SIDE, right=BORDER_SIDE,
                     top=BORDER_SIDE,  bottom=BORDER_SIDE)

MONEY_FMT   = '"$"#,##0'          # formats 250000 as $250,000
CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT        = Alignment(horizontal="left",   vertical="center", wrap_text=False)

# ── Column widths (rough guide) ─────────────────────────────────
col_widths = {
    "A": 22,   # Company_Name
    "B": 28,   # Email
    "C": 16,   # Industry
    "D": 16,   # Account_Owner
    "E": 16,   # Needs_Attention
    "F": 16,   # Payment_Status
    "G": 16,   # Annual_Revenue
    "H": 14,   # Contract_Start
    "I": 20,   # Total_Forecast_Revenue
    "J": 12,   # Q1_Forecast
    "K": 12,   # Q2_Forecast
    "L": 12,   # Q3_Forecast
    "M": 12,   # Q4_Forecast
    "N": 12,   # Confidence
    "O": 14,   # Support_Status
    "P": 16,   # Issue_Type
    "Q": 10,   # Priority
    "R": 14,   # Ticket_Open_Date
    "S": 14,   # Segment
    "T": 18,   # Last_Campaign_Sent
    "U": 12,   # Email_Opens
    "V": 12,   # Subscribed
    "W": 14,   # Phone
}
for col_letter, width in col_widths.items():
    ws.column_dimensions[col_letter].width = width

ws.row_dimensions[1].height = 30   # taller header row

# ── Apply header styling (row 1) ────────────────────────────────
for cell in ws[1]:
    cell.fill      = HEADER_BG
    cell.font      = HEADER_FONT
    cell.alignment = CENTER
    cell.border    = THIN_BORDER

# ── Apply data row styling ───────────────────────────────────────
money_columns = {"G", "I", "J", "K", "L", "M"}   # revenue / forecast cols

for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
    # Alternate row colour for readability
    row_fill = ALT_ROW_BG if row_idx % 2 == 0 else WHITE_BG

    for cell in row:
        col_letter = get_column_letter(cell.column)

        # Base styling
        cell.fill      = row_fill
        cell.border    = THIN_BORDER
        cell.font      = Font(name="Calibri", size=10)
        cell.alignment = CENTER if col_letter in {"E", "F", "G", "I",
                                                   "J", "K", "L", "M",
                                                   "N", "O", "P", "Q",
                                                   "S", "U", "V"} else LEFT

        # Currency format for money columns
        if col_letter in money_columns and cell.value is not None:
            cell.number_format = MONEY_FMT

        # Highlight "Needs Attention" column (E)
        if col_letter == "E":
            cell.font = Font(name="Calibri", size=10, bold=True)
            if cell.value == "Yes":
                cell.fill = FLAG_YES_BG
                cell.font = Font(name="Calibri", size=10, bold=True, color="CC0000")
            elif cell.value == "No":
                cell.fill = FLAG_NO_BG
                cell.font = Font(name="Calibri", size=10, bold=True, color="006600")

# ── Freeze the top row so headers stay visible when scrolling ───
ws.freeze_panes = "A2"

# ── Add auto-filter so users can sort/filter in Excel ───────────
ws.auto_filter.ref = ws.dimensions

# ─────────────────────────────────────────────────────────────────
# STEP 13 – ADD A SUMMARY SHEET
# ─────────────────────────────────────────────────────────────────
print("\nSTEP 10: Adding a Summary sheet...")

ws_sum = wb.create_sheet("Summary")
ws_sum.column_dimensions["A"].width = 32
ws_sum.column_dimensions["B"].width = 20

summary_header_fill = PatternFill("solid", fgColor="1F3864")
summary_header_font = Font(bold=True, color="FFFFFF", size=12, name="Calibri")
summary_value_font  = Font(size=11, name="Calibri")

rows = [
    ("Metric",                              "Value"),
    ("Total Customers",                     len(master)),
    ("Customers Needing Attention",         int((master["Needs_Attention"] == "Yes").sum())),
    ("Total Annual Revenue ($)",            int(master["Annual_Revenue"].sum())),
    ("Total Forecast Revenue ($)",          int(master["Total_Forecast_Revenue"].sum())),
    ("Avg Annual Revenue per Customer ($)", int(master["Annual_Revenue"].mean())),
    ("Overdue Payments",                    int((master["Payment_Status"] == "Overdue").sum())),
    ("Open/In-Progress Support Tickets",   int((master["Support_Status"].isin(["Open", "In Progress"])).sum())),
    ("High Confidence Forecast Customers",  int((master["Confidence"] == "High").sum())),
]

for r_idx, (label, value) in enumerate(rows, start=2):
    cell_a = ws_sum.cell(row=r_idx, column=1, value=label)
    cell_b = ws_sum.cell(row=r_idx, column=2, value=value)

    is_header = r_idx == 2
    for cell in [cell_a, cell_b]:
        cell.border    = THIN_BORDER
        cell.alignment = Alignment(horizontal="left" if cell.column == 1 else "right",
                                   vertical="center")
        if is_header:
            cell.fill = summary_header_fill
            cell.font = summary_header_font
        else:
            cell.font = summary_value_font
            cell.fill = (ALT_ROW_BG if r_idx % 2 == 0 else WHITE_BG)

    # Format currency rows
    if "($)" in label:
        cell_b.number_format = MONEY_FMT

    ws_sum.row_dimensions[r_idx].height = 22

ws_sum.freeze_panes = "A3"

# ─────────────────────────────────────────────────────────────────
# STEP 14 – SAVE
# ─────────────────────────────────────────────────────────────────
wb.save(OUTPUT_FILE)

print(f"\n{'=' * 60}")
print("ALL DONE!")
print(f"  Output file: {OUTPUT_FILE}")
print(f"  Sheet 1: 'Master Customers' – {len(master)} customers, {len(master.columns)} columns")
print(f"  Sheet 2: 'Summary' – key stats at a glance")
print("=" * 60)

# ── Quick preview in the console ────────────────────────────────
print("\nTop 5 customers by Annual Revenue:")
print(master[["Company_Name", "Annual_Revenue", "Total_Forecast_Revenue",
              "Needs_Attention", "Payment_Status", "Support_Status"]].head(5).to_string(index=False))
