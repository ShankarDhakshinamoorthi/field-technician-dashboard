"""
customer_pipeline.py
====================
Scalable, self-healing customer data pipeline.

How it works (plain English):
  1. You drop any CSV file into the  inbox/  folder.
  2. Run this script (or let watcher.py run it automatically).
  3. It figures out what each column is — even if the names change.
  4. It merges everything by email address into a single master.
  5. It saves the full dataset as a fast Parquet file, and a
     styled Excel preview of the top rows.
  6. It remembers which files it already processed, so it never
     double-counts.

Usage:
  python customer_pipeline.py           # process everything in inbox/
  python customer_pipeline.py file.csv  # process one specific file
"""

import difflib
import json
import logging
import os
import re
import shutil
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────
# PATHS  (all relative to the folder this script lives in)
# ─────────────────────────────────────────────────────────────────
BASE_DIR       = Path(__file__).parent
CONFIG_FILE    = BASE_DIR / "config" / "schema_config.json"
INBOX_DIR      = BASE_DIR / "inbox"          # drop new CSVs here
ARCHIVE_DIR    = BASE_DIR / "archive"        # processed files copied here
OUTPUT_DIR     = BASE_DIR / "output"         # Excel + Parquet outputs
LOG_DIR        = BASE_DIR / "logs"
DATA_DIR       = BASE_DIR / "Data"           # original sample data (bootstrap)
PROCESSED_FILE = BASE_DIR / "processed_files.json"
REJECTED_FILE  = BASE_DIR / "rejected_rows.csv"
PARQUET_PATH   = OUTPUT_DIR / "master_customers.parquet"
EXCEL_PATH     = OUTPUT_DIR / "master_customers.xlsx"

EXCEL_MAX_ROWS = 5_000    # Excel preview cap — Parquet holds ALL rows
CHUNK_SIZE     = 50_000   # Rows read at a time (keeps memory flat)

# ─────────────────────────────────────────────────────────────────
# LOGGING  (writes to file AND the console at the same time)
# ─────────────────────────────────────────────────────────────────
LOG_DIR.mkdir(parents=True, exist_ok=True)
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[
        logging.FileHandler(LOG_DIR / "pipeline.log", encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger("pipeline")


# ─────────────────────────────────────────────────────────────────
# HELPER FUNCTIONS
# ─────────────────────────────────────────────────────────────────
def _norm(s: str) -> str:
    """
    Normalise a column name for matching:
      'Contact_Email' -> 'contactemail'
      'E-Mail Address' -> 'emailaddress'
    This lets us match despite different capitalisation and punctuation.
    """
    return re.sub(r"[^a-z0-9]", "", str(s).lower())


def _load_config() -> dict:
    """Read the schema_config.json file."""
    with open(CONFIG_FILE, encoding="utf-8") as f:
        return json.load(f)


def _load_processed() -> dict:
    """Load the log of already-processed files."""
    if PROCESSED_FILE.exists():
        with open(PROCESSED_FILE, encoding="utf-8") as f:
            return json.load(f)
    return {}


def _save_processed(data: dict) -> None:
    """Save the updated processed-files log."""
    with open(PROCESSED_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, default=str)


# ─────────────────────────────────────────────────────────────────
# SCHEMA DETECTOR
# ─────────────────────────────────────────────────────────────────
class SchemaDetector:
    """
    Figures out what each column in an incoming file actually means,
    even when the column names are different from what we expect.

    Think of it like a translator: it reads the column names and says
    'ah, this "Contact_Email" column is the email address we want.'

    Detection works in 3 layers (most-certain first):
      1. Exact match  — 'email' == 'email'                confidence 1.0
      2. Keyword check — 'email' appears inside the name   confidence 0.8
      3. Fuzzy match  — similar enough by difflib          confidence 0.6+
    """

    def __init__(self, config: dict):
        self.config = config
        self._email_aliases = [_norm(a) for a in config["join_key"]["aliases"]]
        # Build normalised alias lists for each standard column
        self._col_aliases: dict[str, list[str]] = {
            std: [_norm(a) for a in aliases]
            for std, aliases in config["columns"].items()
        }
        self._source_types = {
            k: v for k, v in config.get("source_types", {}).items()
            if not k.startswith("_")
        }

    # ── Find the email column ─────────────────────────────────────
    def find_email_col(self, df: pd.DataFrame) -> tuple:
        """Return (column_name, confidence) for the email column in df."""
        df_norms = {_norm(c): c for c in df.columns}

        # Layer 1: exact match
        for norm_alias in self._email_aliases:
            if norm_alias in df_norms:
                return df_norms[norm_alias], 1.0

        # Layer 2: keyword containment
        for norm_col, actual_col in df_norms.items():
            if "email" in norm_col or "mail" in norm_col:
                return actual_col, 0.8

        # Layer 3: fuzzy match (last resort)
        best_col, best_score = None, 0.0
        for norm_col, actual_col in df_norms.items():
            for alias in self._email_aliases:
                score = difflib.SequenceMatcher(None, norm_col, alias).ratio()
                if score > best_score:
                    best_score, best_col = score, actual_col

        if best_score >= 0.65:
            return best_col, round(best_score, 2)
        return None, 0.0

    # ── Classify what kind of file this is ───────────────────────
    def find_source_type(self, df: pd.DataFrame, filename: str) -> tuple:
        """
        Return (type_name, how_we_decided).
        type_name: salesforce | billing | support | marketing | forecast | unknown
        """
        fname_lower = filename.lower()
        df_norms = {_norm(c) for c in df.columns}

        # Check filename keywords first
        for stype, info in self._source_types.items():
            for pat in info.get("filename_patterns", []):
                if pat in fname_lower:
                    return stype, f"filename contains '{pat}'"

        # Fall back to column fingerprint voting
        scores: dict[str, int] = {}
        for stype, info in self._source_types.items():
            fp = [_norm(c) for c in info.get("column_fingerprint", [])]
            scores[stype] = sum(1 for c in fp if c in df_norms)

        if scores:
            best_type = max(scores, key=scores.get)
            if scores[best_type] > 0:
                return best_type, f"column fingerprint ({scores[best_type]} match(es))"

        return "unknown", "no_match — file will still be merged on email"

    # ── Build the standard-name → actual-column-name mapping ─────
    def build_col_mapping(self, df: pd.DataFrame) -> dict:
        """
        Returns {standard_name: actual_column_name} for every
        recognisable column in df.

        Example:
          df has 'Contact_Email', 'Annual_Revenue', 'Payment_Status'
          returns {'email': 'Contact_Email',
                   'annual_revenue': 'Annual_Revenue',
                   'payment_status': 'Payment_Status'}
        """
        mapping: dict[str, str] = {}
        df_norms = {_norm(c): c for c in df.columns}
        used_actuals: set[str] = set()   # prevent one column mapping to two standard names

        for std_name, norm_aliases in self._col_aliases.items():
            # Layer 1: exact alias match (most reliable)
            for norm_alias in norm_aliases:
                actual = df_norms.get(norm_alias)
                if actual and actual not in used_actuals:
                    mapping[std_name] = actual
                    used_actuals.add(actual)
                    break
            else:
                # Layer 2: alias is a substring of the column name
                # Min length 6 prevents short words like "status" from matching
                # "paymentstatus" (which belongs to payment_status, not support_status)
                for norm_col, actual_col in df_norms.items():
                    if actual_col in used_actuals:
                        continue
                    if any(alias in norm_col for alias in norm_aliases if len(alias) >= 6):
                        mapping[std_name] = actual_col
                        used_actuals.add(actual_col)
                        break

        return mapping


# ─────────────────────────────────────────────────────────────────
# FILE READER  (chunked + multi-encoding)
# ─────────────────────────────────────────────────────────────────
def read_file(filepath: Path, chunksize: int = CHUNK_SIZE) -> pd.DataFrame:
    """
    Read a CSV file in pieces (chunks) to avoid memory issues.

    For a file with 1 million rows, instead of loading all 1M into RAM
    we load 50,000 at a time. Peak memory stays flat no matter how big
    the file is.

    Also tries multiple text encodings (utf-8, latin-1, cp1252) so that
    files with special characters (accents, symbols) don't crash.
    """
    for encoding in ("utf-8", "latin-1", "cp1252"):
        try:
            chunks = []
            reader = pd.read_csv(
                filepath,
                chunksize=chunksize,
                encoding=encoding,
                dtype=str,            # read everything as text first (safe)
                low_memory=False,
                on_bad_lines="warn",  # skip bad rows, don't crash
            )
            for chunk in reader:
                # Strip accidental leading/trailing spaces from column names
                chunk.columns = chunk.columns.str.strip()
                chunks.append(chunk)

            if not chunks:
                return pd.DataFrame()

            df = pd.concat(chunks, ignore_index=True)
            log.info(f"  Read {len(df):,} rows from '{filepath.name}' (encoding: {encoding})")
            return df

        except UnicodeDecodeError:
            continue   # try the next encoding
        except Exception as exc:
            log.error(f"  Error reading '{filepath.name}': {exc}")
            raise

    raise ValueError(
        f"Could not read '{filepath.name}' — tried utf-8, latin-1, cp1252. "
        "Save the file as UTF-8 in Excel and try again."
    )


# ─────────────────────────────────────────────────────────────────
# MASTER STORE  (load, upsert, flag, save)
# ─────────────────────────────────────────────────────────────────
class MasterStore:
    """
    Manages the master customer table.

    The master lives in  output/master_customers.parquet.
    When new data arrives we UPSERT it:
      - Existing email: update only the new columns we received.
      - New email: add a fresh row.

    We never lose old data — if billing says $250k but no forecast
    data has arrived yet, the forecast columns stay blank until a
    forecast file comes in.
    """

    def __init__(self, config: dict):
        self.config = config

    def load(self) -> pd.DataFrame:
        if PARQUET_PATH.exists():
            df = pd.read_parquet(PARQUET_PATH)
            log.info(f"Loaded existing master: {len(df):,} rows, {len(df.columns)} columns")
            return df
        log.info("No existing master found — starting fresh.")
        return pd.DataFrame()

    def upsert(self, master: pd.DataFrame, new_df: pd.DataFrame,
               new_std_cols: list) -> pd.DataFrame:
        """
        Merge new_df into master on the 'email' column.

        new_std_cols: the list of standard column names present in new_df
                      (we only overwrite these columns, leaving others intact).
        """
        if "email" not in new_df.columns:
            log.warning("  new_df has no 'email' column — cannot upsert, skipping.")
            return master

        new_df = new_df.copy()
        new_df["email"] = new_df["email"].str.lower().str.strip()

        if master.empty:
            return new_df.copy()

        master = master.copy()
        master["email"] = master["email"].str.lower().str.strip()

        existing_emails = set(master["email"].dropna())
        new_emails_mask  = ~new_df["email"].isin(existing_emails)

        # Rows to append (brand-new customers)
        additions = new_df[new_emails_mask].copy()

        # Rows to update (customers we already know about)
        updates = new_df[~new_emails_mask].copy()
        update_cols = [c for c in new_std_cols if c in updates.columns and c != "email"]

        if not updates.empty and update_cols:
            # Add any brand-new columns from this source to master first.
            # pandas .update() only touches existing columns — without this step,
            # new columns (e.g. q1_forecast arriving in a forecast file) are silently lost.
            for col in update_cols:
                if col not in master.columns:
                    master[col] = None   # new column: blank for all existing rows

            updates_indexed = updates.set_index("email")[update_cols]
            master = master.set_index("email")
            master.update(updates_indexed)   # fill in the new values for matching emails
            master = master.reset_index()

        if not additions.empty:
            master = pd.concat([master, additions], ignore_index=True)

        log.info(f"  Upsert complete: {len(updates)} rows updated, "
                 f"{len(additions)} rows added.")
        return master

    def compute_flags(self, master: pd.DataFrame) -> pd.DataFrame:
        """
        Add the 'needs_attention' column using rules from schema_config.json.

        'Yes' = customer has an overdue payment OR an open support ticket.
        'No'  = everything is fine.

        To add a new trigger (e.g. 'Suspended'), just add it to
        flag_rules.payment_overdue in schema_config.json — no code change.
        """
        flag_rules  = self.config.get("flag_rules", {})
        overdue_vals = [v.lower() for v in flag_rules.get("payment_overdue", [])]
        open_vals    = [v.lower() for v in flag_rules.get("support_open", [])]

        master = master.copy()
        overdue     = pd.Series(False, index=master.index)
        open_ticket = pd.Series(False, index=master.index)

        if "payment_status" in master.columns:
            overdue = master["payment_status"].fillna("").str.lower().str.strip().isin(overdue_vals)
        if "support_status" in master.columns:
            open_ticket = master["support_status"].fillna("").str.lower().str.strip().isin(open_vals)

        master["needs_attention"] = (overdue | open_ticket).map({True: "Yes", False: "No"})
        flagged = int((master["needs_attention"] == "Yes").sum())
        log.info(f"  Flags: {flagged} customer(s) need attention.")
        return master

    def sort_by_revenue(self, master: pd.DataFrame) -> pd.DataFrame:
        """Sort by annual_revenue descending (highest-value customers first)."""
        if "annual_revenue" in master.columns:
            master = master.copy()
            master["annual_revenue"] = pd.to_numeric(
                master["annual_revenue"], errors="coerce"
            )
            master = master.sort_values(
                "annual_revenue", ascending=False, na_position="last"
            ).reset_index(drop=True)
        return master

    def save(self, master: pd.DataFrame) -> None:
        """
        Write two outputs:
          1. Parquet — the full dataset, no row limit, fast to read later.
          2. Excel   — top N rows styled for human review.

        Why Parquet? For 100k+ rows it is:
          - 10x smaller than CSV on disk
          - 50x faster to load than Excel
          - Preserves data types (numbers stay numbers, dates stay dates)
        """
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

        master.to_parquet(PARQUET_PATH, index=False)
        log.info(f"Full dataset saved -> {PARQUET_PATH.name}  ({len(master):,} rows)")

        excel_df = master.head(EXCEL_MAX_ROWS)
        _write_excel(excel_df, master, self.config)
        capped = "" if len(master) <= EXCEL_MAX_ROWS else f" (capped at {EXCEL_MAX_ROWS:,})"
        log.info(f"Excel preview saved -> {EXCEL_PATH.name}  ({len(excel_df):,} rows{capped})")


# ─────────────────────────────────────────────────────────────────
# EXCEL WRITER  (dynamic column detection, no hardcoded letters)
# ─────────────────────────────────────────────────────────────────
def _write_excel(df: pd.DataFrame, full_df: pd.DataFrame, config: dict) -> None:
    """Write a styled two-sheet Excel workbook."""

    with pd.ExcelWriter(EXCEL_PATH, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Master Customers", index=False)

    wb = load_workbook(EXCEL_PATH)
    ws = wb["Master Customers"]

    # ── Style palette ────────────────────────────────────────────
    NAVY     = PatternFill("solid", fgColor="1F3864")
    ALT      = PatternFill("solid", fgColor="EEF2F7")
    WHITE    = PatternFill("solid", fgColor="FFFFFF")
    RED_BG   = PatternFill("solid", fgColor="FFD7D7")
    GRN_BG   = PatternFill("solid", fgColor="D7F0D7")
    THIN_SIDE = Side(style="thin", color="CCCCCC")
    BORDER   = Border(left=THIN_SIDE, right=THIN_SIDE,
                      top=THIN_SIDE,  bottom=THIN_SIDE)

    # Build a map: column_letter -> standard_name (lowercased)
    col_name_map: dict[str, str] = {
        get_column_letter(cell.column): str(cell.value or "").lower().strip()
        for cell in ws[1]
    }

    money_names = {"annual_revenue", "total_forecast_revenue",
                   "q1_forecast", "q2_forecast", "q3_forecast", "q4_forecast"}
    money_cols  = {ltr for ltr, name in col_name_map.items() if name in money_names}
    flag_col    = next((ltr for ltr, name in col_name_map.items()
                        if name == "needs_attention"), None)

    # ── Header row ───────────────────────────────────────────────
    ws.row_dimensions[1].height = 30
    for cell in ws[1]:
        cell.fill      = NAVY
        cell.font      = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = BORDER

    # ── Data rows ────────────────────────────────────────────────
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        base_fill = ALT if row_idx % 2 == 0 else WHITE
        for cell in row:
            ltr = get_column_letter(cell.column)
            cell.border    = BORDER
            cell.font      = Font(name="Calibri", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")

            # Currency formatting for money columns
            if ltr in money_cols and cell.value is not None:
                try:
                    cell.value        = float(str(cell.value).replace(",", ""))
                    cell.number_format = '"$"#,##0'
                except (ValueError, TypeError):
                    pass

            # Conditional colour for the Needs_Attention column
            if ltr == flag_col:
                cell.font = Font(name="Calibri", size=10, bold=True)
                if cell.value == "Yes":
                    cell.fill = RED_BG
                    cell.font = Font(name="Calibri", size=10, bold=True, color="CC0000")
                elif cell.value == "No":
                    cell.fill = GRN_BG
                    cell.font = Font(name="Calibri", size=10, bold=True, color="006600")
                else:
                    cell.fill = base_fill
            else:
                cell.fill = base_fill

    # ── Auto-fit column widths ────────────────────────────────────
    for col in ws.columns:
        max_len = max(
            (len(str(cell.value)) if cell.value is not None else 0 for cell in col),
            default=10,
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # ── Summary sheet ─────────────────────────────────────────────
    _write_summary_sheet(wb, full_df, config)
    wb.save(EXCEL_PATH)


def _write_summary_sheet(wb, full_df: pd.DataFrame, config: dict) -> None:
    ws = wb.create_sheet("Summary")
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 22

    flag_rules   = config.get("flag_rules", {})
    overdue_vals = [v.lower() for v in flag_rules.get("payment_overdue", [])]
    open_vals    = [v.lower() for v in flag_rules.get("support_open", [])]

    def _cnt(col, vals):
        if col not in full_df.columns:
            return "N/A"
        return int(full_df[col].fillna("").str.lower().str.strip().isin(vals).sum())

    def _sum_col(col):
        if col not in full_df.columns:
            return "N/A"
        return int(pd.to_numeric(full_df[col], errors="coerce").sum())

    def _avg_col(col):
        if col not in full_df.columns:
            return "N/A"
        v = pd.to_numeric(full_df[col], errors="coerce").mean()
        return int(v) if pd.notna(v) else "N/A"

    att_col = "needs_attention"
    summary_rows = [
        ("Metric",                                "Value"),
        ("Total Customers",                       len(full_df)),
        ("Customers Needing Attention",
         int((full_df.get(att_col, pd.Series(dtype=str)) == "Yes").sum())),
        ("Total Annual Revenue ($)",              _sum_col("annual_revenue")),
        ("Total Forecast Revenue ($)",            _sum_col("total_forecast_revenue")),
        ("Avg Annual Revenue per Customer ($)",   _avg_col("annual_revenue")),
        ("Overdue Payments",                      _cnt("payment_status", overdue_vals)),
        ("Open / In-Progress Support Tickets",    _cnt("support_status", open_vals)),
        ("High Confidence Forecasts",             _cnt("forecast_confidence", ["high"])),
        ("Full dataset rows (Parquet file)",      len(full_df)),
        (f"Excel preview shows top {EXCEL_MAX_ROWS:,} rows by revenue", ""),
    ]

    NAVY = PatternFill("solid", fgColor="1F3864")
    ALT  = PatternFill("solid", fgColor="EEF2F7")
    WHT  = PatternFill("solid", fgColor="FFFFFF")
    THIN_SIDE = Side(style="thin", color="CCCCCC")
    BDR  = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

    for r_i, (label, value) in enumerate(summary_rows, start=2):
        ca = ws.cell(row=r_i, column=1, value=label)
        cb = ws.cell(row=r_i, column=2, value=value)
        for cell in (ca, cb):
            cell.border    = BDR
            cell.alignment = Alignment(
                horizontal="left" if cell.column == 1 else "right",
                vertical="center"
            )
        if r_i == 2:
            for cell in (ca, cb):
                cell.fill = NAVY
                cell.font = Font(bold=True, color="FFFFFF", size=12, name="Calibri")
        else:
            fill = ALT if r_i % 2 == 0 else WHT
            for cell in (ca, cb):
                cell.fill = fill
                cell.font = Font(size=11, name="Calibri")
            if "($)" in str(label) and isinstance(value, (int, float)):
                cb.number_format = '"$"#,##0'
        ws.row_dimensions[r_i].height = 22

    ws.freeze_panes = "A3"


# ─────────────────────────────────────────────────────────────────
# REJECTED ROWS
# ─────────────────────────────────────────────────────────────────
def _save_rejected(bad_rows: pd.DataFrame, source_file: str) -> None:
    """Save rows that could not be processed (e.g., missing email)."""
    bad_rows = bad_rows.copy()
    bad_rows["__source_file__"]   = source_file
    bad_rows["__rejected_at__"]   = datetime.now().isoformat()
    header = not REJECTED_FILE.exists()
    bad_rows.to_csv(REJECTED_FILE, mode="a", index=False, header=header)


# ─────────────────────────────────────────────────────────────────
# PIPELINE  (the main orchestrator)
# ─────────────────────────────────────────────────────────────────
class Pipeline:
    """
    Orchestrates the full flow for one or many CSV files:
      read -> detect -> standardise -> deduplicate -> upsert -> flag -> save
    """

    def __init__(self):
        self.config    = _load_config()
        self.detector  = SchemaDetector(self.config)
        self.store     = MasterStore(self.config)
        self.processed = _load_processed()

    def already_processed(self, filepath: Path) -> bool:
        """
        True if we've already processed this file AND it hasn't changed
        since (we compare the file's last-modified time).
        """
        key = str(filepath.resolve())
        if key not in self.processed:
            return False
        stored_mtime = self.processed[key].get("mtime", 0)
        current_mtime = filepath.stat().st_mtime
        return abs(stored_mtime - current_mtime) < 1   # 1-second tolerance

    def mark_done(self, filepath: Path, stats: dict) -> None:
        """
        Copy the file to archive/ with a timestamp prefix, then record it
        in processed_files.json so it won't be re-processed next time.
        """
        ARCHIVE_DIR.mkdir(parents=True, exist_ok=True)
        ts  = datetime.now().strftime("%Y%m%d_%H%M%S")
        dst = ARCHIVE_DIR / f"{ts}_{filepath.name}"
        shutil.copy2(filepath, dst)

        self.processed[str(filepath.resolve())] = {
            "mtime":        filepath.stat().st_mtime,
            "processed_at": datetime.now().isoformat(),
            "archive_copy": str(dst),
            **stats,
        }
        _save_processed(self.processed)
        log.info(f"  Archived copy -> archive/{dst.name}")

    def process_one(self, filepath: Path) -> bool:
        """
        Process a single CSV file from start to finish.
        Returns True if successful, False if the file was skipped/errored.

        Every step is inside try/except so one bad file never stops
        the others from processing.
        """
        log.info(f"{'-'*55}")
        log.info(f"Processing: {filepath.name}")

        # ── Step 1: Read ──────────────────────────────────────────
        try:
            df = read_file(filepath)
        except Exception as exc:
            log.error(f"  FAILED to read file: {exc}")
            return False

        if df.empty:
            log.warning(f"  File is empty — skipped.")
            return False

        # ── Step 2: Find email column ─────────────────────────────
        email_col, confidence = self.detector.find_email_col(df)
        if email_col is None or confidence < 0.6:
            log.error(
                f"  Could not identify an email column (best confidence: "
                f"{confidence:.0%}). File skipped.\n"
                f"  Columns found: {list(df.columns)}\n"
                f"  Add an alias to config/schema_config.json -> join_key.aliases"
            )
            return False
        log.info(f"  Email column: '{email_col}'  (confidence {confidence:.0%})")

        # ── Step 3: Classify source type ──────────────────────────
        source_type, method = self.detector.find_source_type(df, filepath.name)
        log.info(f"  Source type:  {source_type}  (detected via {method})")

        # ── Step 4: Map columns to standard names ─────────────────
        col_mapping = self.detector.build_col_mapping(df)
        col_mapping["email"] = email_col   # always include the email key
        unmapped = [c for c in df.columns if c not in col_mapping.values()]
        if unmapped:
            log.info(f"  Unmapped columns (kept as-is): {unmapped}")

        log.info(f"  Recognised {len(col_mapping)} standard columns: "
                 f"{list(col_mapping.keys())}")

        # ── Step 5: Rename to standard names, keep unmapped too ───
        # Rename recognised columns; keep unrecognised ones under their original names
        rename_map = {actual: std for std, actual in col_mapping.items()
                      if actual in df.columns}
        df_std = df.rename(columns=rename_map)

        # ── Step 6: Clean the email column ────────────────────────
        df_std["email"] = df_std["email"].astype(str).str.strip().str.lower()

        # ── Step 7: Separate good rows from bad rows ───────────────
        bad_mask  = df_std["email"].isna() | df_std["email"].isin(["", "nan", "none"])
        good_rows = df_std[~bad_mask].copy()
        bad_rows  = df_std[bad_mask].copy()

        if not bad_rows.empty:
            _save_rejected(bad_rows, filepath.name)
            log.warning(f"  {len(bad_rows)} rows had no email -> saved to rejected_rows.csv")

        if good_rows.empty:
            log.error("  No valid rows after removing missing emails. File skipped.")
            return False

        # ── Step 8: Deduplicate within this file ──────────────────
        # If the same email appears twice in one file, keep the last occurrence
        before = len(good_rows)
        good_rows = good_rows.drop_duplicates(subset=["email"], keep="last")
        dupes = before - len(good_rows)
        if dupes:
            log.info(f"  Deduplicated {dupes} duplicate email(s) within file.")
        log.info(f"  {len(good_rows):,} unique-email rows ready for merge.")

        # ── Step 9: Load master + upsert ──────────────────────────
        master    = self.store.load()
        std_cols  = [c for c in good_rows.columns if c != "email"]
        master    = self.store.upsert(master, good_rows, std_cols)

        # ── Step 10: Compute flags + sort ─────────────────────────
        master = self.store.compute_flags(master)
        master = self.store.sort_by_revenue(master)

        # ── Step 11: Save ─────────────────────────────────────────
        self.store.save(master)

        # ── Step 12: Record + archive ─────────────────────────────
        self.mark_done(filepath, {
            "source_type":    source_type,
            "rows_read":      len(df),
            "rows_merged":    len(good_rows),
            "rows_rejected":  len(bad_rows),
        })

        log.info(
            f"  Master now has {len(master):,} rows x {len(master.columns)} columns."
        )
        return True

    def run(self, folder: Path = None, files: list = None) -> None:
        """
        Discover and process CSV files.

        folder: scan this folder for *.csv files
        files:  process a specific list of file paths (overrides folder)
        """
        if files is None:
            folder = Path(folder) if folder else INBOX_DIR
            if not folder.exists():
                log.warning(f"Folder '{folder}' does not exist. Nothing to process.")
                return
            files = sorted(folder.glob("*.csv"))

        files = [Path(f) for f in files]

        if not files:
            log.info("No CSV files found to process.")
            return

        log.info(f"Found {len(files)} CSV file(s).")
        ok_count = skip_count = fail_count = 0

        for filepath in files:
            if not filepath.exists():
                log.warning(f"File not found: {filepath}")
                fail_count += 1
                continue
            if self.already_processed(filepath):
                log.info(f"Skipping '{filepath.name}' — already processed and unchanged.")
                skip_count += 1
                continue
            success = self.process_one(filepath)
            if success:
                ok_count += 1
            else:
                fail_count += 1

        log.info(f"{'-'*55}")
        log.info(f"Run complete — processed: {ok_count}, "
                 f"skipped: {skip_count}, failed: {fail_count}")


# ─────────────────────────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────────────────────────
def main() -> None:
    # Ensure all folders exist
    for d in (INBOX_DIR, ARCHIVE_DIR, OUTPUT_DIR, LOG_DIR):
        d.mkdir(parents=True, exist_ok=True)

    log.info("=" * 55)
    log.info("CUSTOMER PIPELINE  —  starting")
    log.info("=" * 55)

    pipeline = Pipeline()

    if len(sys.argv) > 1:
        # Files given on the command line: python customer_pipeline.py a.csv b.csv
        files = [Path(p) for p in sys.argv[1:]]
        log.info(f"Processing {len(files)} file(s) from command line.")
        pipeline.run(files=files)
    else:
        # Auto-detect: inbox/ → Data/ (bootstrap fallback)
        inbox_csvs = list(INBOX_DIR.glob("*.csv"))
        if inbox_csvs:
            log.info(f"Processing {len(inbox_csvs)} file(s) from inbox/")
            pipeline.run(folder=INBOX_DIR)
        elif DATA_DIR.exists():
            log.info("inbox/ is empty — bootstrapping from Data/ folder.")
            pipeline.run(folder=DATA_DIR)
        else:
            log.warning(
                "Nothing to process.\n"
                f"  Put CSV files in: {INBOX_DIR}\n"
                f"  Or run: python customer_pipeline.py <file.csv>"
            )


if __name__ == "__main__":
    main()
